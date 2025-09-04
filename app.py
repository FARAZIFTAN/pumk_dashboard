from flask import Flask, render_template, request, jsonify, redirect, url_for, session, flash, send_file
from flask_mysqldb import MySQL
import MySQLdb.cursors
import config
from decimal import Decimal
from functools import wraps
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.config.from_object(config.Config)
app.secret_key = 'your-secret-key-here'  # Ganti dengan secret key yang aman
mysql = MySQL(app)

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def get_summary_stats():
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    
    # Get total realisasi and outstanding
    cursor.execute("""
        SELECT 
            COUNT(*) as total_mitra,
            SUM(Jumlah_Realisasi_Rp) as total_realisasi,
            SUM(Outstanding_Rp) as total_outstanding
        FROM mitra_binaan
    """)
    stats = cursor.fetchone()
    
    # Get kolektabilitas distribution
    cursor.execute("""
        SELECT 
            Kolektabilitas_BUMN,
            COUNT(*) as count
        FROM mitra_binaan
        GROUP BY Kolektabilitas_BUMN
    """)
    kolektabilitas = cursor.fetchall()
    
    return stats, kolektabilitas

@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
        
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        remember = request.form.get('remember_me')
        
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute('SELECT * FROM users WHERE email = %s AND password = %s', (email, password))
        user = cursor.fetchone()
        
        if user:
            session['user_id'] = user['id']
            session['name'] = user['name']
            session.permanent = True if remember else False
            return redirect(url_for('dashboard'))
        else:
            flash('Email atau password salah!')
    
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
        
    if request.method == 'POST':
        name = f"{request.form['first_name']} {request.form['last_name']}"
        email = request.form['email']
        password = request.form['password']
        confirm_password = request.form['confirm_password']
        terms = request.form.get('terms')
        
        if not terms:
            flash('Anda harus menyetujui Syarat dan Ketentuan!')
            return render_template('register.html')
            
        if password != confirm_password:
            flash('Password tidak cocok!')
            return render_template('register.html')
        
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        try:
            # Check if email already exists
            cursor.execute('SELECT * FROM users WHERE email = %s', (email,))
            if cursor.fetchone():
                flash('Email sudah terdaftar!')
                return render_template('register.html')
                
            cursor.execute('INSERT INTO users (name, email, password) VALUES (%s, %s, %s)',
                         (name, email, password))
            mysql.connection.commit()
            flash('Registrasi berhasil! Silakan login.')
            return redirect(url_for('login'))
        except Exception as e:
            print(f"Error during registration: {str(e)}")
            flash('Terjadi kesalahan saat mendaftar. Silakan coba lagi.')
            return render_template('register.html')
        finally:
            cursor.close()
    
    return render_template('register.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/export_excel')
def export_excel():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    try:
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

        # Ambil data mitra dengan semua kolom
        query = """
            SELECT
                No,
                Regional_Office,
                Kantor_Cabang,
                Kode_Unit,
                BRI_Unit,
                Nama_Mitra_Binaan,
                Alamat,
                Provinsi,
                Kabupaten_Kota,
                Sektor_Ekonomi_BUMN,
                Jangka_Waktu,
                Tanggal_Realisasi,
                Tanggal_Jatuh_Tempo,
                Jumlah_Realisasi_Rp,
                Outstanding_Rp,
                Kolektabilitas_BUMN,
                Tanggal_Lunas,
                Jenis_Usaha,
                Accrued_Interest,
                Angsuran_Bunga_yang_diterima_Giro_Bulanan
            FROM mitra_binaan
            ORDER BY No
        """
        cursor.execute(query)
        data = cursor.fetchall()

        # Buat Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Data PUMK"

        # Tulis header dengan semua kolom
        headers = [
            'No', 'Regional Office', 'Kantor Cabang', 'Kode Unit', 'BRI Unit',
            'Nama Mitra Binaan', 'Alamat', 'Provinsi', 'Kabupaten/Kota',
            'Sektor Ekonomi BUMN', 'Jangka Waktu', 'Tanggal Realisasi',
            'Tanggal Jatuh Tempo', 'Jumlah Realisasi (Rp)', 'Outstanding (Rp)',
            'Kolektabilitas BUMN', 'Tanggal Lunas', 'Jenis Usaha',
            'Accrued Interest', 'Angsuran Bunga yang diterima Giro Bulanan'
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Style header
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Tulis data
        for row, item in enumerate(data, 2):
            ws.cell(row=row, column=1, value=item['No'])
            ws.cell(row=row, column=2, value=item['Regional_Office'])
            ws.cell(row=row, column=3, value=item['Kantor_Cabang'])
            ws.cell(row=row, column=4, value=item['Kode_Unit'])
            ws.cell(row=row, column=5, value=item['BRI_Unit'])
            ws.cell(row=row, column=6, value=item['Nama_Mitra_Binaan'])
            ws.cell(row=row, column=7, value=item['Alamat'])
            ws.cell(row=row, column=8, value=item['Provinsi'])
            ws.cell(row=row, column=9, value=item['Kabupaten_Kota'])
            ws.cell(row=row, column=10, value=item['Sektor_Ekonomi_BUMN'])
            ws.cell(row=row, column=11, value=item['Jangka_Waktu'])
            ws.cell(row=row, column=12, value=item['Tanggal_Realisasi'])
            ws.cell(row=row, column=13, value=item['Tanggal_Jatuh_Tempo'])
            ws.cell(row=row, column=14, value=item['Jumlah_Realisasi_Rp'])
            ws.cell(row=row, column=15, value=item['Outstanding_Rp'])
            ws.cell(row=row, column=16, value=item['Kolektabilitas_BUMN'])
            ws.cell(row=row, column=17, value=item['Tanggal_Lunas'])
            ws.cell(row=row, column=18, value=item['Jenis_Usaha'])
            ws.cell(row=row, column=19, value=item['Accrued_Interest'])
            ws.cell(row=row, column=20, value=item['Angsuran_Bunga_yang_diterima_Giro_Bulanan'])

        # Format currency columns
        currency_format = '#,##0'
        for row in range(2, len(data) + 2):
            ws.cell(row=row, column=14).number_format = currency_format  # Jumlah Realisasi
            ws.cell(row=row, column=15).number_format = currency_format  # Outstanding
            ws.cell(row=row, column=19).number_format = currency_format  # Accrued Interest
            ws.cell(row=row, column=20).number_format = currency_format  # Angsuran Bunga

        # Format date columns
        date_format = 'DD/MM/YYYY'
        for row in range(2, len(data) + 2):
            if ws.cell(row=row, column=12).value:  # Tanggal Realisasi
                ws.cell(row=row, column=12).number_format = date_format
            if ws.cell(row=row, column=13).value:  # Tanggal Jatuh Tempo
                ws.cell(row=row, column=13).number_format = date_format
            if ws.cell(row=row, column=17).value:  # Tanggal Lunas
                ws.cell(row=row, column=17).number_format = date_format

        # Adjust column widths
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

        # Create response
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'PUMK_Data_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )

    except Exception as e:
        print(f"Error exporting Excel: {str(e)}")
        flash('Terjadi kesalahan saat mengexport data', 'error')
        return redirect(url_for('dashboard'))

    finally:
        cursor.close()

@app.route('/import_excel', methods=['POST'])
@login_required
def import_excel():
    if 'file' not in request.files:
        flash('No file selected', 'error')
        return redirect(url_for('dashboard'))

    file = request.files['file']
    if file.filename == '':
        flash('No file selected', 'error')
        return redirect(url_for('dashboard'))

    if not file.filename.endswith('.xlsx'):
        flash('Please upload an Excel file (.xlsx)', 'error')
        return redirect(url_for('dashboard'))

    try:
        wb = load_workbook(file)
        ws = wb.active
        cursor = mysql.connection.cursor()

        imported_count = 0
        # Skip header row
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None or row[0] == '':  # Skip empty rows
                continue

            # Ensure numeric fields are handled
            jumlah_realisasi = float(row[5]) if row[5] else 0
            outstanding = float(row[6]) if row[6] else 0

            cursor.execute("""
                INSERT INTO mitra_binaan (No, Nama_Mitra_Binaan, Alamat, Provinsi, Kabupaten_Kota, Jumlah_Realisasi_Rp, Outstanding_Rp, Kolektabilitas_BUMN)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE
                Nama_Mitra_Binaan=VALUES(Nama_Mitra_Binaan),
                Alamat=VALUES(Alamat),
                Provinsi=VALUES(Provinsi),
                Kabupaten_Kota=VALUES(Kabupaten_Kota),
                Jumlah_Realisasi_Rp=VALUES(Jumlah_Realisasi_Rp),
                Outstanding_Rp=VALUES(Outstanding_Rp),
                Kolektabilitas_BUMN=VALUES(Kolektabilitas_BUMN)
            """, (row[0], row[1], row[2], row[3], row[4], jumlah_realisasi, outstanding, row[7]))
            imported_count += 1

        mysql.connection.commit()
        flash(f'Successfully imported {imported_count} records', 'success')

    except Exception as e:
        print(f"Import error: {str(e)}")
        flash('Error importing data. Please check the file format.', 'error')

    finally:
        cursor.close()

    return redirect(url_for('dashboard'))

@app.route('/')
@login_required
def index():
    return redirect(url_for('dashboard'))

@app.route('/dashboard')
@login_required
def dashboard():
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # Get base data with all columns
    query = """
        SELECT
            No,
            Regional_Office,
            Kantor_Cabang,
            Kode_Unit,
            BRI_Unit,
            Nama_Mitra_Binaan,
            Alamat,
            Provinsi,
            Kabupaten_Kota,
            Sektor_Ekonomi_BUMN,
            Jangka_Waktu,
            Tanggal_Realisasi,
            Tanggal_Jatuh_Tempo,
            Jumlah_Realisasi_Rp,
            Outstanding_Rp,
            Kolektabilitas_BUMN,
            Tanggal_Lunas,
            Jenis_Usaha,
            Accrued_Interest,
            Angsuran_Bunga_yang_diterima_Giro_Bulanan
        FROM mitra_binaan
        ORDER BY No
        LIMIT 1000
    """
    cursor.execute(query)
    data = cursor.fetchall()

    # Get summary statistics
    stats, kolektabilitas = get_summary_stats()

    return render_template('dashboard.html',
                         mitra=data,
                         stats=stats,
                         kolektabilitas=kolektabilitas)

@app.route('/mitra-binaan')
@login_required
def mitra_binaan():
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cursor.execute('SELECT * FROM mitra_binaan ORDER BY No')
    mitra = cursor.fetchall()
    return render_template('mitra_binaan.html', mitra=mitra)

@app.route('/mitra-binaan/add', methods=['GET', 'POST'])
@login_required
def add_mitra_binaan():
    if request.method == 'POST':
        form = request.form
        try:
            cursor = mysql.connection.cursor()
            cursor.execute("""
                INSERT INTO mitra_binaan (
                    Regional_Office, Kantor_Cabang, Kode_Unit, BRI_Unit, Nama_Mitra_Binaan,
                    Alamat, Provinsi, Kabupaten_Kota, Sektor_Ekonomi_BUMN, Jangka_Waktu,
                    Tanggal_Realisasi, Tanggal_Jatuh_Tempo, Jumlah_Realisasi_Rp, Outstanding_Rp,
                    Kolektabilitas_BUMN, Tanggal_Lunas, Jenis_Usaha, Accrued_Interest,
                    Angsuran_Bunga_yang_diterima_Giro_Bulanan
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                form.get('Regional_Office'), form.get('Kantor_Cabang'), form.get('Kode_Unit'), form.get('BRI_Unit'), form.get('Nama_Mitra_Binaan'),
                form.get('Alamat'), form.get('Provinsi'), form.get('Kabupaten_Kota'), form.get('Sektor_Ekonomi_BUMN'), form.get('Jangka_Waktu'),
                form.get('Tanggal_Realisasi'), form.get('Tanggal_Jatuh_Tempo'), form.get('Jumlah_Realisasi_Rp'), form.get('Outstanding_Rp'),
                form.get('Kolektabilitas_BUMN'), form.get('Tanggal_Lunas'), form.get('Jenis_Usaha'), form.get('Accrued_Interest'),
                form.get('Angsuran_Bunga_yang_diterima_Giro_Bulanan')
            ))
            mysql.connection.commit()
            flash('Data mitra binaan berhasil ditambahkan.', 'success')
            return redirect(url_for('mitra_binaan'))
        except Exception as e:
            flash(f'Gagal menambahkan data: {str(e)}', 'danger')
            return redirect(url_for('mitra_binaan'))
    return render_template('mitra_binaan_form.html', action='add')

@app.route('/mitra-binaan/edit/<int:no>', methods=['GET', 'POST'])
@login_required
def edit_mitra_binaan(no):
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    if request.method == 'POST':
        form = request.form
        try:
            cursor.execute("""
                UPDATE mitra_binaan SET
                    Regional_Office=%s, Kantor_Cabang=%s, Kode_Unit=%s, BRI_Unit=%s, Nama_Mitra_Binaan=%s,
                    Alamat=%s, Provinsi=%s, Kabupaten_Kota=%s, Sektor_Ekonomi_BUMN=%s, Jangka_Waktu=%s,
                    Tanggal_Realisasi=%s, Tanggal_Jatuh_Tempo=%s, Jumlah_Realisasi_Rp=%s, Outstanding_Rp=%s,
                    Kolektabilitas_BUMN=%s, Tanggal_Lunas=%s, Jenis_Usaha=%s, Accrued_Interest=%s,
                    Angsuran_Bunga_yang_diterima_Giro_Bulanan=%s
                WHERE No=%s
            """, (
                form.get('Regional_Office'), form.get('Kantor_Cabang'), form.get('Kode_Unit'), form.get('BRI_Unit'), form.get('Nama_Mitra_Binaan'),
                form.get('Alamat'), form.get('Provinsi'), form.get('Kabupaten_Kota'), form.get('Sektor_Ekonomi_BUMN'), form.get('Jangka_Waktu'),
                form.get('Tanggal_Realisasi'), form.get('Tanggal_Jatuh_Tempo'), form.get('Jumlah_Realisasi_Rp'), form.get('Outstanding_Rp'),
                form.get('Kolektabilitas_BUMN'), form.get('Tanggal_Lunas'), form.get('Jenis_Usaha'), form.get('Accrued_Interest'),
                form.get('Angsuran_Bunga_yang_diterima_Giro_Bulanan'), no
            ))
            mysql.connection.commit()
            flash('Data mitra binaan berhasil diperbarui.', 'success')
            return redirect(url_for('mitra_binaan'))
        except Exception as e:
            flash(f'Gagal memperbarui data: {str(e)}', 'danger')
            return redirect(url_for('mitra_binaan'))
    else:
        cursor.execute('SELECT * FROM mitra_binaan WHERE No = %s', (no,))
        mitra = cursor.fetchone()
        if mitra is None:
            flash('Data mitra binaan tidak ditemukan.', 'warning')
            return redirect(url_for('mitra_binaan'))
        return render_template('mitra_binaan_form.html', mitra=mitra, action='edit')

@app.route('/mitra-binaan/delete/<int:no>', methods=['POST'])
@login_required
def delete_mitra_binaan(no):
    try:
        cursor = mysql.connection.cursor()
        cursor.execute('DELETE FROM mitra_binaan WHERE No = %s', (no,))
        mysql.connection.commit()
        flash('Data mitra binaan berhasil dihapus.', 'success')
    except Exception as e:
        flash(f'Gagal menghapus data: {str(e)}', 'danger')
    return redirect(url_for('mitra_binaan'))

@app.route('/keuangan')
@login_required
def keuangan():
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    # Get summary statistics
    stats, kolektabilitas = get_summary_stats()
    return render_template('keuangan.html', stats=stats, kolektabilitas=kolektabilitas)

@app.route('/laporan')
@login_required
def laporan():
    return render_template('laporan.html')

@app.route('/api/filter')
def filter_data():
    provinsi = request.args.get('provinsi')
    kolektabilitas = request.args.get('kolektabilitas')
    
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    
    query = "SELECT * FROM mitra_binaan WHERE 1=1"
    if provinsi:
        query += f" AND Provinsi = '{provinsi}'"
    if kolektabilitas:
        query += f" AND Kolektabilitas_BUMN = '{kolektabilitas}'"
    
    cursor.execute(query)
    data = cursor.fetchall()
    return jsonify(data)

if __name__ == '__main__':
    app.run(debug=True)
